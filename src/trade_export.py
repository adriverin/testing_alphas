import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def extract_trades_from_backtest(strategy_returns, portfolio_info, price_data, alpha_name="Strategy"):
    """
    Extract individual trades from backtest results.
    
    FIXED VERSION: Now properly calculates trade P&L to match actual portfolio returns
    by accounting for how the backtesting system calculates daily returns.
    
    Args:
        strategy_returns: Daily strategy returns
        portfolio_info: DataFrame with weights, turnover data
        price_data: DataFrame with asset price and return data
        alpha_name: Name of the strategy/alpha
        
    Returns:
        pd.DataFrame: Detailed trade records
    """
    trades = []
    
    # Get asset return data (this is what the backtesting system uses)
    if 'returns' not in price_data.columns:
        print("⚠️ No return data available to calculate trade P&L")
        return pd.DataFrame()
    
    asset_returns = price_data['returns']
    
    # Get position weights
    weights = portfolio_info['weights']
    
    # Process each asset separately
    for asset in weights.index.get_level_values('asset').unique():
        asset_weights = weights.xs(asset, level='asset')
        asset_returns_series = asset_returns.xs(asset, level='asset')
        
        # Align weights and returns
        common_dates = asset_weights.index.intersection(asset_returns_series.index)
        asset_weights = asset_weights.reindex(common_dates).fillna(0)
        asset_returns_series = asset_returns_series.reindex(common_dates).fillna(0)
        
        # Track current position
        current_position = 0.0
        entry_date = None
        entry_weight = None
        cumulative_trade_return = 0.0
        daily_contributions = []
        
        for date in common_dates:
            new_weight = asset_weights[date]
            new_position = 1 if new_weight > 0 else (-1 if new_weight < 0 else 0)
            daily_return = asset_returns_series[date]
            
            # If we have a position, accumulate the daily contribution
            if current_position != 0 and entry_date is not None:
                # This is the key fix: calculate daily contribution exactly as backtesting system does
                daily_contribution = new_weight * daily_return
                daily_contributions.append((date, daily_contribution))
                cumulative_trade_return += daily_contribution
            
            # Check for position changes
            if new_position != current_position:
                # Close existing position if any
                if current_position != 0 and entry_date is not None:
                    exit_date = date
                    
                    # Calculate trade metrics based on actual daily contributions
                    days_held = (exit_date - entry_date).days
                    
                    # Total return contribution from this trade
                    total_contribution = cumulative_trade_return
                    
                    # Calculate equivalent percentage return for display
                    # This is approximate since we're dealing with weighted returns
                    if abs(entry_weight) > 0:
                        equivalent_pnl_percent = (total_contribution / abs(entry_weight)) * 100
                    else:
                        equivalent_pnl_percent = 0.0
                    
                    trade_type = "LONG" if current_position == 1 else "SHORT"
                    
                    # Get entry and exit prices for reference (approximate)
                    entry_idx = common_dates.get_loc(entry_date)
                    exit_idx = common_dates.get_loc(exit_date)
                    
                    # Calculate approximate prices from returns
                    if 'close' in price_data.columns:
                        prices = price_data['close'].xs(asset, level='asset')
                        entry_price = prices.reindex(common_dates).fillna(method='ffill').iloc[entry_idx]
                        exit_price = prices.reindex(common_dates).fillna(method='ffill').iloc[exit_idx]
                    else:
                        # Approximate from cumulative returns
                        cumulative_returns = (1 + asset_returns_series).cumprod()
                        entry_price = cumulative_returns.iloc[entry_idx] * 100  # Normalize to ~100
                        exit_price = cumulative_returns.iloc[exit_idx] * 100
                    
                    trades.append({
                        'Asset': asset,
                        'Strategy': alpha_name,
                        'Trade_Type': trade_type,
                        'Entry_Date': entry_date,
                        'Exit_Date': exit_date,
                        'Entry_Price': entry_price,
                        'Exit_Price': exit_price,
                        'Entry_Weight': entry_weight,
                        'Days_Held': days_held,
                        'PnL_Percent': equivalent_pnl_percent,
                        'Weight_Impact': total_contribution,  # This is the actual portfolio contribution
                        'Trade_Result': 'WIN' if total_contribution > 0 else 'LOSS',
                        'Daily_Contributions': len(daily_contributions)
                    })
                
                # Reset for new position
                daily_contributions = []
                cumulative_trade_return = 0.0
                
                # Open new position if not flat
                if new_position != 0:
                    current_position = new_position
                    entry_date = date
                    entry_weight = new_weight
                else:
                    current_position = 0
                    entry_date = None
                    entry_weight = None
        
        # Handle any remaining open position
        if current_position != 0 and entry_date is not None:
            exit_date = common_dates[-1]
            days_held = (exit_date - entry_date).days
            
            # Calculate final contribution
            total_contribution = cumulative_trade_return
            
            if abs(entry_weight) > 0:
                equivalent_pnl_percent = (total_contribution / abs(entry_weight)) * 100
            else:
                equivalent_pnl_percent = 0.0
            
            trade_type = "LONG" if current_position == 1 else "SHORT"
            
            # Get approximate prices
            entry_idx = common_dates.get_loc(entry_date)
            exit_idx = len(common_dates) - 1
            
            if 'close' in price_data.columns:
                prices = price_data['close'].xs(asset, level='asset')
                entry_price = prices.reindex(common_dates).fillna(method='ffill').iloc[entry_idx]
                exit_price = prices.reindex(common_dates).fillna(method='ffill').iloc[exit_idx]
            else:
                cumulative_returns = (1 + asset_returns_series).cumprod()
                entry_price = cumulative_returns.iloc[entry_idx] * 100
                exit_price = cumulative_returns.iloc[exit_idx] * 100
            
            trades.append({
                'Asset': asset,
                'Strategy': alpha_name,
                'Trade_Type': trade_type,
                'Entry_Date': entry_date,
                'Exit_Date': exit_date,
                'Entry_Price': entry_price,
                'Exit_Price': exit_price,
                'Entry_Weight': entry_weight,
                'Days_Held': days_held,
                'PnL_Percent': equivalent_pnl_percent,
                'Weight_Impact': total_contribution,
                'Trade_Result': 'WIN' if total_contribution > 0 else 'LOSS',
                'Status': 'OPEN',
                'Daily_Contributions': len(daily_contributions)
            })
    
    if not trades:
        print("⚠️ No trades found in backtest data")
        return pd.DataFrame()
    
    # Convert to DataFrame and sort by entry date
    trades_df = pd.DataFrame(trades)
    trades_df = trades_df.sort_values('Entry_Date').reset_index(drop=True)
    
    # Validation: Check that weight impacts sum to approximately the total portfolio return
    total_weight_impact = trades_df['Weight_Impact'].sum()
    total_strategy_return = strategy_returns.sum()
    
    print(f"📊 Trade validation:")
    print(f"   Sum of trade impacts: {total_weight_impact:.4f}")
    print(f"   Total strategy returns: {total_strategy_return:.4f}")
    print(f"   Difference: {abs(total_weight_impact - total_strategy_return):.4f}")
    
    if abs(total_weight_impact - total_strategy_return) > 0.01:
        print("⚠️  Warning: Large discrepancy detected between trade impacts and strategy returns")
    
    return trades_df

def calculate_trade_statistics(trades_df):
    """Calculate comprehensive trade statistics"""
    if trades_df.empty:
        return {}
    
    total_trades = len(trades_df)
    winning_trades = len(trades_df[trades_df['PnL_Percent'] > 0])
    losing_trades = len(trades_df[trades_df['PnL_Percent'] < 0])
    
    win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
    
    avg_win = trades_df[trades_df['PnL_Percent'] > 0]['PnL_Percent'].mean() if winning_trades > 0 else 0
    avg_loss = trades_df[trades_df['PnL_Percent'] < 0]['PnL_Percent'].mean() if losing_trades > 0 else 0
    
    profit_factor = abs(avg_win * winning_trades / (avg_loss * losing_trades)) if avg_loss != 0 and losing_trades > 0 else np.inf
    
    avg_days_held = trades_df['Days_Held'].mean()
    max_win = trades_df['PnL_Percent'].max()
    max_loss = trades_df['PnL_Percent'].min()
    
    # Calculate by trade type
    long_trades = trades_df[trades_df['Trade_Type'] == 'LONG']
    short_trades = trades_df[trades_df['Trade_Type'] == 'SHORT']
    
    # Portfolio return calculations
    total_weight_impact = trades_df['Weight_Impact'].sum()
    
    # Calculate compounded return from weight impacts
    # This is an approximation since we don't have exact daily timing
    daily_avg_impact = total_weight_impact / len(trades_df) if len(trades_df) > 0 else 0
    
    stats = {
        'Total_Trades': total_trades,
        'Winning_Trades': winning_trades,
        'Losing_Trades': losing_trades,
        'Win_Rate_Percent': win_rate,
        'Average_Win_Percent': avg_win,
        'Average_Loss_Percent': avg_loss,
        'Profit_Factor': profit_factor,
        'Average_Days_Held': avg_days_held,
        'Best_Trade_Percent': max_win,
        'Worst_Trade_Percent': max_loss,
        'Long_Trades': len(long_trades),
        'Long_Win_Rate': (len(long_trades[long_trades['PnL_Percent'] > 0]) / len(long_trades) * 100) if len(long_trades) > 0 else 0,
        'Short_Trades': len(short_trades),
        'Short_Win_Rate': (len(short_trades[short_trades['PnL_Percent'] > 0]) / len(short_trades) * 100) if len(short_trades) > 0 else 0,
        'Total_PnL_Percent': trades_df['PnL_Percent'].sum(),
        'Total_Weight_Impact': total_weight_impact,
        'Portfolio_Return_Sum': total_weight_impact,  # Sum of daily returns
        'Portfolio_Return_Approx_Compound': (1 + total_weight_impact) - 1 if total_weight_impact > -1 else total_weight_impact  # Approximation
    }
    
    return stats

def export_trades_to_csv(trades_df, stats, output_path, alpha_name="Strategy"):
    """
    Export trades and statistics to CSV files.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Main trades file
    trades_path = output_path.with_suffix('.csv')
    trades_df.to_csv(trades_path, index=False)
    
    # Statistics file
    stats_path = output_path.with_name(f"{output_path.stem}_statistics.csv")
    stats_df = pd.DataFrame(list(stats.items()), columns=['Metric', 'Value'])
    stats_df.to_csv(stats_path, index=False)
    
    # Winning trades
    winning_trades = trades_df[trades_df['PnL_Percent'] > 0]
    if not winning_trades.empty:
        wins_path = output_path.with_name(f"{output_path.stem}_winning_trades.csv")
        winning_trades.to_csv(wins_path, index=False)
    
    # Losing trades
    losing_trades = trades_df[trades_df['PnL_Percent'] < 0]
    if not losing_trades.empty:
        losses_path = output_path.with_name(f"{output_path.stem}_losing_trades.csv")
        losing_trades.to_csv(losses_path, index=False)
    
    print(f"📊 Trade analysis exported to: {trades_path.parent}")
    return trades_path

def export_trades_to_excel(trades_df, stats, output_path, alpha_name="Strategy"):
    """
    Export trades and statistics to a formatted Excel file.
    """
    if not EXCEL_AVAILABLE:
        print("⚠️ openpyxl not available, falling back to CSV export")
        return export_trades_to_csv(trades_df, stats, output_path, alpha_name)
    
    output_path = Path(output_path).with_suffix('.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Summary formatting
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    
    # Add title
    ws_summary['A1'] = f"{alpha_name} - Trading Results Summary"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:B1')
    
    # Add timestamp
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A2'].font = Font(size=10, italic=True)
    
    # Add statistics
    row = 4
    for key, value in stats.items():
        ws_summary[f'A{row}'] = key.replace('_', ' ').title()
        ws_summary[f'A{row}'].font = header_font
        
        if isinstance(value, (int, float)):
            if 'Percent' in key or 'Rate' in key:
                ws_summary[f'B{row}'] = f"{value:.2f}%"
            elif 'Factor' in key:
                ws_summary[f'B{row}'] = f"{value:.2f}" if value != np.inf else "∞"
            elif 'Days' in key:
                ws_summary[f'B{row}'] = f"{value:.1f}"
            else:
                ws_summary[f'B{row}'] = int(value) if value == int(value) else f"{value:.2f}"
        else:
            ws_summary[f'B{row}'] = str(value)
        
        row += 1
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Create Trades sheet
    if not trades_df.empty:
        ws_trades = wb.create_sheet("All Trades")
        
        # Add headers
        headers = list(trades_df.columns)
        for i, header in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=i)
            cell.value = header.replace('_', ' ')
            cell.font = header_font
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add data
        for row_idx, (_, row) in enumerate(trades_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_trades.cell(row=row_idx, column=col_idx)
                
                if isinstance(value, pd.Timestamp):
                    cell.value = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    if 'Percent' in headers[col_idx-1] or 'Weight' in headers[col_idx-1]:
                        cell.value = f"{value:.2f}%"
                    elif 'Price' in headers[col_idx-1]:
                        cell.value = f"${value:.2f}"
                    else:
                        cell.value = value
                else:
                    cell.value = str(value) if value is not None else ""
                
                # Color coding for P&L
                if 'PnL' in headers[col_idx-1]:
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        if value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths for trades sheet
        for column in ws_trades.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_trades.column_dimensions[column_letter].width = adjusted_width
        
        # Create filtered views for wins/losses
        winning_trades = trades_df[trades_df['PnL_Percent'] > 0]
        losing_trades = trades_df[trades_df['PnL_Percent'] < 0]
        
        if not winning_trades.empty:
            ws_wins = wb.create_sheet("Winning Trades")
            # Copy headers
            for i, header in enumerate(headers, 1):
                cell = ws_wins.cell(row=1, column=i)
                cell.value = header.replace('_', ' ')
                cell.font = header_font
                cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Copy winning trades data
            for row_idx, (_, row) in enumerate(winning_trades.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_wins.cell(row=row_idx, column=col_idx)
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        if 'Percent' in headers[col_idx-1] or 'Weight' in headers[col_idx-1]:
                            cell.value = f"{value:.2f}%"
                        elif 'Price' in headers[col_idx-1]:
                            cell.value = f"${value:.2f}"
                        else:
                            cell.value = value
                    else:
                        cell.value = str(value) if value is not None else ""
        
        if not losing_trades.empty:
            ws_losses = wb.create_sheet("Losing Trades")
            # Copy headers
            for i, header in enumerate(headers, 1):
                cell = ws_losses.cell(row=1, column=i)
                cell.value = header.replace('_', ' ')
                cell.font = header_font
                cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Copy losing trades data
            for row_idx, (_, row) in enumerate(losing_trades.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_losses.cell(row=row_idx, column=col_idx)
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        if 'Percent' in headers[col_idx-1] or 'Weight' in headers[col_idx-1]:
                            cell.value = f"{value:.2f}%"
                        elif 'Price' in headers[col_idx-1]:
                            cell.value = f"${value:.2f}"
                        else:
                            cell.value = value
                    else:
                        cell.value = str(value) if value is not None else ""
    
    # Save workbook
    wb.save(output_path)
    print(f"📊 Trade analysis exported to: {output_path}")
    return output_path

def export_backtest_trades(alpha_calculator, price_data, alpha_name="alpha998", output_dir="export_trades_to_csv/trade_exports", export_format="excel", stop_loss_pct=None):
    """
    Main function to export trades from a specific alpha's backtest.
    
    Args:
        alpha_calculator: Alpha101 instance
        price_data: Price data DataFrame
        alpha_name: Name of alpha to analyze (default: alpha998)
        output_dir: Directory to save CSV files
        export_format: Export format ("excel" or "csv")
        stop_loss_pct: Optional stop-loss percentage (e.g., -5.0 for 5% loss)
    
    Returns:
        Path to exported CSV file
    """
    from src.backtests import run_rank_backtest
    
    print(f"🔍 Extracting trades for {alpha_name}...")
    if stop_loss_pct is not None:
        print(f"🛡️ Individual position stop-loss enabled: {stop_loss_pct}%")
    
    # Get alpha signals
    if not hasattr(alpha_calculator, alpha_name):
        print(f"❌ Alpha {alpha_name} not found")
        return None
    
    alpha_series = getattr(alpha_calculator, alpha_name)().dropna()
    
    if alpha_series.empty:
        print(f"❌ No signals generated for {alpha_name}")
        return None
    
    # Run backtest with optional stop-loss
    strategy_returns, portfolio_info = run_rank_backtest(price_data, alpha_series, stop_loss_pct)
    
    if strategy_returns.empty:
        print(f"❌ No returns generated for {alpha_name}")
        return None
    
    # Extract trades
    trades_df = extract_trades_from_backtest(strategy_returns, portfolio_info, price_data, alpha_name)
    
    if trades_df.empty:
        print(f"❌ No trades extracted for {alpha_name}")
        return None
    
    # Add stop-loss information to trades
    if hasattr(portfolio_info, 'attrs') and 'stopped_positions' in portfolio_info.attrs:
        stopped_positions = portfolio_info.attrs['stopped_positions']
        
        # Mark trades that were stopped out
        trades_df['Stop_Loss_Exit'] = False
        for stopped_pos in stopped_positions:
            mask = (
                (trades_df['Asset'] == stopped_pos['asset']) &
                (trades_df['Entry_Date'] == stopped_pos['entry_date']) &
                (trades_df['Exit_Date'] == stopped_pos['stop_date'])
            )
            trades_df.loc[mask, 'Stop_Loss_Exit'] = True
            trades_df.loc[mask, 'Trade_Result'] = 'STOP_LOSS'
    else:
        trades_df['Stop_Loss_Exit'] = False
    
    # Calculate statistics
    stats = calculate_trade_statistics(trades_df)
    
    # Add stop-loss statistics
    if hasattr(portfolio_info, 'attrs'):
        stats['Stop_Loss_Triggers'] = portfolio_info.attrs.get('stop_loss_triggers', 0)
        stats['Stop_Loss_Percentage'] = stop_loss_pct if stop_loss_pct is not None else 'Disabled'
        
        # Calculate stop-loss specific metrics
        stop_loss_trades = trades_df[trades_df['Stop_Loss_Exit'] == True]
        stats['Stop_Loss_Count'] = len(stop_loss_trades)
        if len(stop_loss_trades) > 0:
            stats['Avg_Stop_Loss_Days'] = stop_loss_trades['Days_Held'].mean()
            stats['Avg_Stop_Loss_PnL'] = stop_loss_trades['PnL_Percent'].mean()
        else:
            stats['Avg_Stop_Loss_Days'] = 0
            stats['Avg_Stop_Loss_PnL'] = 0
    
    # Export based on format
    output_path = Path(output_dir) / f"{alpha_name}_trades_{datetime.now().strftime('%Y%m%d_%H%M')}"
    if export_format.lower() == "excel":
        export_path = export_trades_to_excel(trades_df, stats, output_path, alpha_name)
    else:
        export_path = export_trades_to_csv(trades_df, stats, output_path, alpha_name)
    
    # Calculate actual portfolio returns for comparison
    sum_of_daily_returns = strategy_returns.sum()
    compounded_return = (1 + strategy_returns).prod() - 1
    
    # Print summary
    print(f"\n📈 {alpha_name} Trade Summary:")
    print(f"   Total Trades: {stats['Total_Trades']}")
    print(f"   Win Rate: {stats['Win_Rate_Percent']:.1f}%")
    print(f"   Profit Factor: {stats['Profit_Factor']:.2f}")
    print(f"   Average Days Held: {stats['Average_Days_Held']:.1f}")
    print(f"   Best Trade: {stats['Best_Trade_Percent']:.2f}%")
    print(f"   Worst Trade: {stats['Worst_Trade_Percent']:.2f}%")
    
    # Print stop-loss summary
    if stop_loss_pct is not None:
        print(f"\n🛡️ Stop-Loss Summary:")
        print(f"   Stop-Loss Threshold: {stop_loss_pct}%")
        print(f"   Positions Stopped Out: {stats['Stop_Loss_Count']}")
        if stats['Stop_Loss_Count'] > 0:
            print(f"   Avg Days to Stop-Loss: {stats['Avg_Stop_Loss_Days']:.1f}")
            print(f"   Avg Stop-Loss P&L: {stats['Avg_Stop_Loss_PnL']:.2f}%")
    
    print(f"\n📊 Portfolio Return Analysis:")
    print(f"   Sum of Trade Impacts: {stats['Total_Weight_Impact']:.4f}")
    print(f"   Actual Sum of Daily Returns: {sum_of_daily_returns:.4f}")
    print(f"   Difference: {abs(stats['Total_Weight_Impact'] - sum_of_daily_returns):.4f}")
    print(f"   Actual Compounded Return: {compounded_return:.4f}")
    print(f"   Compounding Effect: {compounded_return - sum_of_daily_returns:.4f}")
    
    if abs(stats['Total_Weight_Impact'] - sum_of_daily_returns) < 0.1:
        print(f"   ✅ Trade impacts correctly match sum of daily returns")
    else:
        print(f"   ⚠️  Large discrepancy detected - check calculation")
    
    return export_path 